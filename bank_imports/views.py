from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from GARCA.utils import add_breadcrumb, clear_breadcrumbs
from accounts.models import Account, AccountKeyword
from async_tasks.tasks import recalculate_balances_after_date
from entries.models import Entry
from transactions.models import Transaction
from .forms import  BankImportForm
import io
import csv
from decimal import Decimal, InvalidOperation 
from datetime import datetime, timedelta 
from django.contrib import messages
from django.db import transaction
from django.views import View
from django.utils.decorators import method_decorator
import os 
import logging 
import openpyxl
import xlrd 

logger = logging.getLogger(__name__)

class ImportMovementsMixin:
    @transaction.atomic
    def import_movements(self, account, import_data):
        count = 0
        affected_accounts = set()
        min_date = None
        affected_accounts.add(account.id)

        for movement in import_data:
            counterpart_account = None
            for keyword in AccountKeyword.objects.all():
                if keyword.keyword.lower() in movement['description'].lower():
                    counterpart_account = keyword.account
                    break

            entry = Entry.objects.create(
                date=movement['date'],
                description=movement['description']
            )

            amount = movement['amount']
            debit = amount if amount > 0 else 0
            credit = abs(amount) if amount < 0 else 0

            Transaction.objects.create(
                entry=entry,
                account=account,
                debit=debit,
                credit=credit
            )

            if counterpart_account:
                affected_accounts.add(counterpart_account.id)
                Transaction.objects.create(
                    entry=entry,
                    account=counterpart_account,
                    debit=credit, # Counterpart transaction is reversed
                    credit=debit  # Counterpart transaction is reversed
                )

            if min_date is None or movement['date'] < min_date:
                min_date = movement['date']

            count += 1

        for account_id in affected_accounts:
            recalculate_balances_after_date.delay(
                min_date,
                account_id
            )

        return count

class BankImportView(ImportMovementsMixin, View):
    template_name = 'import_movements.html'

    def get(self, request):
        clear_breadcrumbs(request)
        add_breadcrumb(request, 'Importar movimientos' , request.path)
        form = BankImportForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = BankImportForm(request.POST, request.FILES)
        if form.is_valid():
            bank_provider = form.cleaned_data['bank_provider']
            account = form.cleaned_data['account']
            file = request.FILES['file']
            
            bank_provider_display_name = dict(form.fields['bank_provider'].choices).get(bank_provider, bank_provider)

            preview_data = None
            import_data = None

            _, file_ext_from_name = os.path.splitext(file.name)
            file_extension_lower = file_ext_from_name.lower()

            provider_specific_allowed_extensions = {
                'ing': ['.xlsx', '.xls'], # ING: Soporta archivos Excel
                'edenred': ['.csv', '.txt', '.xlsx'],
                'paypal': ['.csv'],
            }

            current_allowed_extensions = provider_specific_allowed_extensions.get(bank_provider)

            if not current_allowed_extensions:
                messages.error(request, f"Proveedor bancario '{bank_provider_display_name}' no tiene configuración de extensiones permitidas.")
                return render(request, self.template_name, {'form': form})

            if file_extension_lower not in current_allowed_extensions:
                expected_extensions_str = ", ".join(current_allowed_extensions)
                messages.error(request, f"Para {bank_provider_display_name}, el archivo debe ser de tipo {expected_extensions_str}. Archivo proporcionado: {file.name}")
                return render(request, self.template_name, {'form': form})
            
            try:
                if bank_provider == 'ing':
                    preview_data, import_data = BankImportView.process_ing_file(file)
                    if not preview_data and not import_data:
                        messages.error(request, "Error al procesar el archivo ING. Verifique que sea un archivo Excel (.xlsx o .xls) válido y que el formato de las columnas sea el esperado.")
                        return render(request, self.template_name, {'form': form})               
                elif bank_provider == 'edenred':
                    preview_data, import_data = BankImportView.process_edenred_file(file)
                    if not preview_data and not import_data: 
                        messages.error(request, "Error al procesar el archivo Edenred. Verifique el formato y la codificación del archivo (UTF-8 o Latin-1 para TXT/CSV, o que sea un XLSX válido).")
                        return render(request, self.template_name, {'form': form})
                elif bank_provider == 'paypal':
                    preview_data, import_data = BankImportView.process_paypal_file(file)
                    if not preview_data and not import_data:
                        messages.error(request, "Error al procesar el archivo PayPal. Verifique que sea un CSV válido (UTF-8 o Latin-1) y que contenga transacciones 'Completado'.")
                        return render(request, self.template_name, {'form': form})
                else:
                    messages.error(request, f"Proveedor bancario '{bank_provider_display_name}' no soportado para la importación.")
                    return render(request, self.template_name, {'form': form})
            except Exception as e: 
                messages.error(request, f"Error inesperado al procesar el archivo para {bank_provider_display_name}: {e}")
                return render(request, self.template_name, {'form': form})

            if 'import' in request.POST and import_data:
                try:
                    duplicates = []
                    non_duplicates = []
                    
                    for movement in import_data:
                        if self.is_duplicate(account, movement):
                            duplicates.append(movement)
                        else:
                            non_duplicates.append(movement)
                    
                    if non_duplicates:
                        success_count = self.import_movements(account, non_duplicates)
                        if duplicates:
                            messages.success(request, f'Se han importado {success_count} movimientos, pero se han encontrado algunos movimientos duplicados')
                        else:
                            messages.success(request, f'Se han importado {success_count} movimientos')
                    
                    if duplicates:
                        session_duplicates = []
                        for movement in duplicates:
                            movement_copy = movement.copy()
                            movement_copy['date'] = movement['date'].strftime('%Y-%m-%d')
                            movement_copy['amount'] = str(movement['amount'])
                            session_duplicates.append(movement_copy)
                        
                        request.session['pending_import'] = {
                            'account_id': account.id,
                            'duplicates': session_duplicates
                        }
                        return redirect('bank_import_duplicates')
                    return redirect('bank_import')
                except Exception as e:
                    messages.error(request, f"Error durante el proceso de importación o verificación de duplicados: {e}")
                    return render(request, self.template_name, {'form': form})

        return render(request, self.template_name, {'form': form})

    def is_duplicate(self, account, movement):
        debit_amount = movement['amount'] if movement['amount'] > 0 else 0
        credit_amount = abs(movement['amount']) if movement['amount'] < 0 else 0
        
        return Transaction.objects.filter(
            account=account,
            entry__date=movement['date'],
            debit=debit_amount,
            credit=credit_amount
        ).exists()

    @staticmethod
    def process_ing_file(file):
        """
        Dispatcher for ING file processing. Determines if it's .xls or .xlsx
        and calls the appropriate handler.
        """
        file_name = file.name.lower()
        if file_name.endswith('.xlsx'):
            logger.info(f"Processing ING file {file.name} as XLSX.")
            return BankImportView._process_ing_xlsx(file)
        elif file_name.endswith('.xls'):
            logger.info(f"Processing ING file {file.name} as XLS.")
            return BankImportView._process_ing_xls(file)
        else:
            logger.error(f"Unsupported file extension for ING: {file.name}")
            return [], []

    @staticmethod
    def _process_ing_xlsx(file_obj): # Renamed from process_ing_file, handles .xlsx
        preview_data = []
        import_data = []
        try:            
            file_obj.seek(0) # Asegurar que el puntero del archivo esté al inicio
            workbook = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = workbook.active
            logger.info("Archivo ING Excel abierto correctamente.")
        except Exception as e: # Broad exception for issues like non-Excel file, corrupted file
            logger.error(f"Error al abrir el archivo ING Excel: {e}")
            return [], []
        header_found = False

        data_start_actual_row = 0 

        logger.debug("Iniciando búsqueda de cabecera en archivo ING XLSX.")
        expected_header_texts = ["F. VALOR", "CATEGORÍA", "DESCRIPCIÓN", "IMPORTE (€)"]
        expected_header_signature_normalized = [h.strip().lower() for h in expected_header_texts]

        for i, row_tuple in enumerate(sheet.iter_rows(min_row=1, max_row=10)): 
            row_values_str_normalized = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in row_tuple]
            
            if len(row_values_str_normalized) >= 7 and \
               row_values_str_normalized[0] == expected_header_signature_normalized[0] and \
               row_values_str_normalized[1] == expected_header_signature_normalized[1] and \
               row_values_str_normalized[3] == expected_header_signature_normalized[2] and \
               row_values_str_normalized[6] == expected_header_signature_normalized[3]:
                header_found = True
                logger.info(f"Cabecera ING (XLSX) encontrada en la fila {i+1}.")                
                data_start_actual_row = (i + 1) + 1 
                break
            else:
                logger.debug(f"Fila {i+1} (XLSX) no coincide con la cabecera ING esperada. Valores normalizados: {row_values_str_normalized[:7]}")
        
        if not header_found:
            logger.warning("Cabecera ING (XLSX) no encontrada.")
            return [], []

        for row_cells_tuple in sheet.iter_rows(min_row=data_start_actual_row):
            row_values = [cell.value for cell in row_cells_tuple]

            if not row_values or row_values[0] is None: # Saltar filas vacías o sin fecha
                logger.debug(f"Fila de datos (XLSX) saltada por estar vacía o sin fecha: {row_values}")
                continue

            try:
                date_val = row_values[0] # F. VALOR (Column A)
                category = str(row_values[1]) if row_values[1] is not None else "" # CATEGORÍA (Column B)
                subcategory = str(row_values[2]) if row_values[2] is not None else "" # SUBCATEGORÍA (Column C)
                description = str(row_values[3]) if row_values[3] is not None else "" # DESCRIPCIÓN (Column D)
                comment = str(row_values[4]) if row_values[4] is not None else "" # COMENTARIO (Column E)
                amount_val = row_values[6] # IMPORTE (€) (Column G)

                parsed_date = None
                if isinstance(date_val, datetime): # openpyxl might parse it as datetime
                    parsed_date = date_val.date()
                elif isinstance(date_val, str):
                    try:
                        parsed_date = datetime.strptime(date_val.strip(), '%d/%m/%Y').date()
                    except ValueError:
                        logger.warning(f"Error al parsear fecha string (XLSX) '{date_val}' en formato dd/mm/yyyy. Saltando fila. Datos: {row_values}")                        
                        continue 
                elif isinstance(date_val, (int, float)): # Sometimes dates are Excel serial numbers
                    try:
                        logger.debug(f"Intentando parsear fecha Excel (XLSX) (número de serie): {date_val}")
                        parsed_date = (datetime(1899, 12, 30) + timedelta(days=date_val)).date()
                    except Exception as e_date_serial:
                        logger.warning(f"Error al convertir fecha Excel (XLSX) (número de serie) '{date_val}': {e_date_serial}. Saltando fila. Datos: {row_values}")
                        continue
                else:
                    logger.warning(f"Formato de fecha (XLSX) no reconocido o nulo: '{date_val}' (tipo: {type(date_val)}). Saltando fila. Datos: {row_values}")
                    continue

                amount_decimal = None
                if amount_val is not None:
                    logger.debug(f"Procesando importe (XLSX): '{amount_val}'")
                    if isinstance(amount_val, (float, int)):
                        amount_decimal = Decimal(str(amount_val))
                    else: # Asumir string, manejar formato europeo
                        amount_str = str(amount_val).replace('.', '').replace(',', '.')
                        amount_decimal = Decimal(amount_str)
                else:
                    logger.warning(f"Importe (XLSX) nulo. Saltando fila. Datos: {row_values}")
                    continue # Skip if amount is missing
                
                logger.debug(f"Fila (XLSX) procesada: Fecha={parsed_date}, Importe={amount_decimal}, Desc={description[:30]}")
                
                preview_data.append({
                    'date': parsed_date,
                    'category': f"{category} - {subcategory}".strip(" -"),
                    'description': description,
                    'amount': amount_decimal
                })
                import_data.append({
                    'date': parsed_date,
                    'category': category,
                    'subcategory': subcategory,
                    'description': description,
                    'comment': comment,
                    'amount': amount_decimal
                })
            except (ValueError, TypeError, IndexError, InvalidOperation) as e:
                logger.warning(f"Skipping row during ING XLSX import due to error: {e}. Row data: {row_values}")
                pass
        logger.info(f"Procesamiento de archivo ING XLSX finalizado. Movimientos para vista previa: {len(preview_data)}, Movimientos para importar: {len(import_data)}")
        return preview_data, import_data

    @staticmethod
    def _process_ing_xls(file_obj): # Handles .xls files using xlrd
        preview_data = []
        import_data = []
        try:
            file_obj.seek(0)
            workbook = xlrd.open_workbook(file_contents=file_obj.read())
            sheet = workbook.sheet_by_index(0) # Asumir la primera hoja
            logger.info("Archivo ING XLS (Excel 97-2003) abierto correctamente con xlrd.")
        except xlrd.XLRDError as e:
            logger.error(f"Error al abrir el archivo ING XLS con xlrd: {e}")
            return [], []
        except Exception as e: # Other potential errors
            logger.error(f"Error inesperado al procesar archivo ING XLS con xlrd: {e}")
            return [], []

        logger.debug("Iniciando búsqueda de cabecera en archivo ING XLS.")
        header_found = False
        data_start_actual_row = 0

        expected_header_texts = ["F. VALOR", "CATEGORÍA", "DESCRIPCIÓN", "IMPORTE (€)"]
        expected_header_signature_normalized = [h.strip().lower() for h in expected_header_texts]

        for i in range(min(sheet.nrows, 10)):
            try:
                row_tuple_values = sheet.row_values(i) # type: ignore
            except IndexError: # Should not happen if i < sheet.nrows
                continue

            row_values_str_normalized = [str(cell_value).strip().lower() if cell_value is not None else "" for cell_value in row_tuple_values]

            if len(row_values_str_normalized) >= 7 and \
               row_values_str_normalized[0] == expected_header_signature_normalized[0] and \
               row_values_str_normalized[1] == expected_header_signature_normalized[1] and \
               row_values_str_normalized[3] == expected_header_signature_normalized[2] and \
               row_values_str_normalized[6] == expected_header_signature_normalized[3]:
                header_found = True
                logger.info(f"Cabecera ING (XLS) encontrada en la fila {i+1}.")
                data_start_actual_row = i + 1
                break
            else:
                logger.debug(f"Fila {i+1} (XLS) no coincide con la cabecera ING esperada. Valores normalizados: {row_values_str_normalized[:7]}")

        if not header_found:
            logger.warning("Cabecera ING (XLS) no encontrada.")
            return [], []

        for row_idx in range(data_start_actual_row, sheet.nrows):
            try:
                row_values = sheet.row_values(row_idx) # type: ignore
            except IndexError:
                continue

            if not row_values or row_values[0] is None or str(row_values[0]).strip() == "":
                logger.debug(f"Fila de datos (XLS) saltada por estar vacía o sin fecha: {row_values}")
                continue

            try:
                date_val_raw = row_values[0]
                category = str(row_values[1]) if row_values[1] is not None else ""
                subcategory = str(row_values[2]) if row_values[2] is not None else ""
                description = str(row_values[3]) if row_values[3] is not None else ""
                comment = str(row_values[4]) if row_values[4] is not None else ""
                amount_val = row_values[6]

                parsed_date = None
                cell_type_date = sheet.cell_type(row_idx, 0) # Column 0 for date

                if cell_type_date == xlrd.XL_CELL_DATE:
                    try:
                        dt_obj = xlrd.xldate_as_datetime(date_val_raw, workbook.datemode)
                        parsed_date = dt_obj.date()
                    except Exception as e_date:
                        logger.warning(f"Error al convertir fecha xlrd (XL_CELL_DATE) '{date_val_raw}': {e_date}. Saltando fila. Datos: {row_values}")
                        continue
                elif cell_type_date == xlrd.XL_CELL_TEXT:
                    try:
                        parsed_date = datetime.strptime(str(date_val_raw).strip(), '%d/%m/%Y').date()
                    except ValueError:
                        logger.warning(f"Error al parsear fecha string (XLS) '{date_val_raw}' en formato dd/mm/yyyy. Saltando fila. Datos: {row_values}")
                        continue
                elif cell_type_date == xlrd.XL_CELL_NUMBER: # Podría ser una fecha almacenada como número
                    try:
                        dt_obj = xlrd.xldate_as_datetime(date_val_raw, workbook.datemode)
                        parsed_date = dt_obj.date()
                        logger.debug(f"Fecha (XLS) parseada de XL_CELL_NUMBER: {date_val_raw} -> {parsed_date}")
                    except Exception: # If not a valid date serial, try string parse if it looks like one
                        try:
                            parsed_date = datetime.strptime(str(date_val_raw).strip(), '%d/%m/%Y').date()
                        except ValueError:
                            logger.warning(f"Formato de fecha (XLS) XL_CELL_NUMBER no reconocido como fecha serial o string dd/mm/yyyy: '{date_val_raw}'. Saltando fila. Datos: {row_values}")
                            continue
                else:
                    logger.warning(f"Formato de fecha (XLS) no reconocido o nulo (tipo xlrd: {cell_type_date}, valor: '{date_val_raw}'). Saltando fila. Datos: {row_values}")
                    continue

                amount_decimal = None
                if amount_val is not None:
                    logger.debug(f"Procesando importe (XLS): '{amount_val}' (tipo: {type(amount_val)})")
                    try:
                        if isinstance(amount_val, (float, int)):
                            amount_decimal = Decimal(str(amount_val))
                        else: # Asumir string, manejar formato europeo
                            amount_str = str(amount_val).replace('.', '').replace(',', '.')
                            amount_decimal = Decimal(amount_str)
                    except InvalidOperation as e_amount:
                        logger.warning(f"Error al convertir importe (XLS) '{amount_val}' a Decimal: {e_amount}. Saltando fila. Datos: {row_values}")
                        continue
                else:
                    logger.warning(f"Importe (XLS) nulo. Saltando fila. Datos: {row_values}")
                    continue
                
                logger.debug(f"Fila (XLS) procesada: Fecha={parsed_date}, Importe={amount_decimal}, Desc={description[:30]}")
                
                preview_data.append({
                    'date': parsed_date,
                    'category': f"{category} - {subcategory}".strip(" -"),
                    'description': description,
                    'amount': amount_decimal
                })
                import_data.append({
                    'date': parsed_date,
                    'category': category,
                    'subcategory': subcategory,
                    'description': description,
                    'comment': comment,
                    'amount': amount_decimal
                })
            except Exception as e:
                logger.warning(f"Skipping row during ING XLS import due to unexpected error: {e}. Row data: {row_values}")
                pass
        logger.info(f"Procesamiento de archivo ING XLS finalizado. Movimientos para vista previa: {len(preview_data)}, Movimientos para importar: {len(import_data)}")
        return preview_data, import_data

    @staticmethod
    def _parse_edenred_row(row_parts):
        """Helper para parsear una fila de datos de Edenred (común a TXT y XLSX)."""
        if len(row_parts) >= 3: # type: ignore
            date_str = str(row_parts[0]).strip() # Asegurar que es string para .xlsx
            description = str(row_parts[1]).strip()
            amount_str = str(row_parts[2]).strip()

            if not date_str and not description and not amount_str:
                return None
            
            parsed_date = None
            try:
                if isinstance(row_parts[0], datetime): # Si ya es datetime (de openpyxl)
                    parsed_date = row_parts[0].date()
                elif ' ' in date_str: 
                    parsed_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').date()
                else: 
                    parsed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return None

            amount_decimal = None
            try:
                amount_decimal = Decimal(amount_str.replace(',', '.'))
            except InvalidOperation:
                return None
            
            final_amount = amount_decimal if "RECARGA" in description.upper() else -amount_decimal

            return {
                'date': parsed_date,
                'description': description,
                'amount': final_amount
            }
        return None

    @staticmethod
    def _process_edenred_text(file):
        preview_data = []
        import_data = []
        content = None
        try:
            content = file.read().decode('utf-8')
        except UnicodeDecodeError:
            file.seek(0) 
            try:
                content = file.read().decode('latin-1')
            except UnicodeDecodeError:
                return [], [] 

        data_io = io.StringIO(content)
        lines = data_io.readlines()
        
        data_start_index = -1
        expected_header_cols = ["Fecha", "Detalle movimiento", "Importe"]

        for i, line_content in enumerate(lines):
            cols = [col.strip() for col in line_content.split('\t')]
            if len(cols) >= len(expected_header_cols) and \
               cols[:len(expected_header_cols)] == expected_header_cols:
                data_start_index = i + 1
                break
        
        if data_start_index == -1:
            return [], []

        for i in range(data_start_index, len(lines)):
            line = lines[i].strip()
            if not line: 
                continue
            row_parts = [part.strip() for part in line.split('\t')]
            parsed_movement = BankImportView._parse_edenred_row(row_parts)
            if parsed_movement:
                preview_data.append(parsed_movement)
                import_data.append(parsed_movement)
        return preview_data, import_data

    @staticmethod
    def _process_edenred_xlsx(file):
        preview_data = []
        import_data = []
        try:
            file.seek(0)
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active
        except Exception as e:
            return [], []

        data_start_index = -1
        expected_header_cols = ["Fecha", "Detalle movimiento", "Importe"]

        for i, row in enumerate(sheet.iter_rows()):
            cols = [str(cell.value).strip() if cell.value is not None else "" for cell in row[:len(expected_header_cols)]]
            if cols == expected_header_cols:
                data_start_index = i + 1
                break
        if data_start_index == -1:
            return [], []
        for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=data_start_index + 1)): # +1 porque min_row es 1-based
            # Extraer valores de las celdas. openpyxl puede devolver directamente tipos como datetime.
            row_values = [cell.value for cell in row_cells]
            
            # Verificar si la fila está esencialmente vacía para evitar procesar filas en blanco al final
            if all(value is None or str(value).strip() == "" for value in row_values[:3]): # Chequear las primeras 3 columnas
                continue

            parsed_movement = BankImportView._parse_edenred_row(row_values)
            if parsed_movement:
                preview_data.append(parsed_movement)
                import_data.append(parsed_movement)
                
        return preview_data, import_data

    @staticmethod
    def process_edenred_file(file):
        """Procesa un archivo de Edenred (TXT, CSV delimitado por tabs, o XLSX)."""
        file_name_lower = file.name.lower()
        if file_name_lower.endswith('.xlsx'):
            return BankImportView._process_edenred_xlsx(file)
        elif file_name_lower.endswith(('.txt', '.csv')): # Asumir que CSV es como TXT para Edenred
            return BankImportView._process_edenred_text(file)
        return [], [] # Tipo de archivo no soportado para Edenred

    @staticmethod
    def process_paypal_file(file):
        preview_data = []
        import_data = []
        
        try:
            file.seek(0) # type: ignore
            content = file.read().decode('utf-8-sig') # Handles UTF-8 with BOM
        except UnicodeDecodeError:
            file.seek(0) # type: ignore
            try:
                content = file.read().decode('latin-1') # Fallback encoding
            except UnicodeDecodeError:
                return [], [] 

        csv_file_like = io.StringIO(content)
        csv_reader = csv.reader(csv_file_like)
        try:
            header = next(csv_reader)
        except StopIteration:
            return [], []
        for row_number, row in enumerate(csv_reader, start=2):
            if len(row) < 8: # PayPal CSV tiene "Importe" en el índice 7
                continue

            try: # type: ignore
                date_str = row[0].strip()
                name_description = row[3].strip()  # "Nombre" field
                status = row[5].strip()
                amount_str = row[7].strip()  # "Importe" field

                if status.lower() != "completado":
                    continue # Process only completed transactions
                if not date_str or not name_description or not amount_str:
                    continue # Skip rows with missing essential data
                parsed_date = datetime.strptime(date_str, '%d/%m/%Y').date()
                amount = Decimal(amount_str.replace(',', '.')) # PayPal uses comma as decimal separator
                movement_data = {
                    'date': parsed_date,
                    'description': name_description,
                    'amount': amount,
                    'category': '',
                    'subcategory': '',
                    'comment': ''
                }
                preview_data.append(movement_data)
                import_data.append(movement_data)
            except (ValueError, InvalidOperation):
                pass 
        return preview_data, import_data

@method_decorator(csrf_exempt, name='dispatch')
class BankImportPreviewView(View):
    def post(self, request):
        form = BankImportForm(request.POST, request.FILES) 
        if form.is_valid():
            bank_provider = form.cleaned_data['bank_provider']
            file = request.FILES['file'] 
            bank_provider_display_name = dict(form.fields['bank_provider'].choices).get(bank_provider, bank_provider)
            
            preview_data = None            

            _, file_ext_from_name = os.path.splitext(file.name)
            file_extension_lower = file_ext_from_name.lower()

            provider_specific_allowed_extensions = { # type: ignore
                'ing': ['.xlsx', '.xls'], # Changed to support Excel files for ING
                'edenred': ['.csv', '.txt', '.xlsx'],
                'paypal': ['.csv'],
            }
            current_allowed_extensions = provider_specific_allowed_extensions.get(bank_provider)

            if not current_allowed_extensions:
                 return JsonResponse({'success': False, 'error': f"Proveedor bancario '{bank_provider_display_name}' no tiene configuración de extensiones."})

            if file_extension_lower not in current_allowed_extensions:
                expected_extensions_str = ", ".join(current_allowed_extensions)
                return JsonResponse({'success': False, 'error': f"Para {bank_provider_display_name}, el archivo debe ser de tipo {expected_extensions_str}. Archivo: {file.name}"})

            try:
                if bank_provider == 'ing':
                    preview_data, _ = BankImportView.process_ing_file(file)
                    if not preview_data and not _:
                         return JsonResponse({'success': False, 'error': 'Error al procesar el archivo ING. Verifique que sea un archivo Excel (.xlsx o .xls) válido y el formato de columnas.'})
                elif bank_provider == 'edenred':
                    preview_data, _ = BankImportView.process_edenred_file(file)
                    if not preview_data and not _:
                         return JsonResponse({'success': False, 'error': 'Error al procesar el archivo Edenred. Verifique el formato y codificación.'})
                elif bank_provider == 'paypal':
                    preview_data, _ = BankImportView.process_paypal_file(file)
                    if not preview_data and not _:
                        return JsonResponse({'success': False, 'error': "Error al procesar el archivo PayPal. Verifique que sea un CSV válido y contenga transacciones 'Completado'."})
                else:
                    return JsonResponse({'success': False, 'error': f"Proveedor bancario '{bank_provider_display_name}' no soportado para vista previa."})
                
                if preview_data is not None:
                    serializable_preview_data = []
                    for item in preview_data:
                        item_copy = item.copy()
                        item_copy['date'] = item_copy['date'].strftime('%Y-%m-%d')
                        item_copy['amount'] = str(item_copy['amount'])
                        serializable_preview_data.append(item_copy)
                    return JsonResponse({'success': True, 'preview_data': serializable_preview_data}) # type: ignore
                else:
                    return JsonResponse({'success': False, 'error': 'No se pudieron generar datos de vista previa.'})

            except Exception as e: 
                return JsonResponse({'success': False, 'error': f'Error al procesar el archivo para vista previa ({bank_provider_display_name}): {e}'})
        
        errors = form.errors.as_json()
        return JsonResponse({'success': False, 'errors': errors})

class BankImportDuplicatesView(ImportMovementsMixin, View):
    template_name = 'import_duplicates.html'

    def get(self, request):
        pending_import = request.session.get('pending_import')
        if not pending_import:
            return redirect('bank_import')

        duplicates_for_template = []
        for movement_str_data in pending_import['duplicates']:
            movement_copy = movement_str_data.copy()
            movement_copy['date'] = datetime.strptime(movement_str_data['date'], '%Y-%m-%d').date()
            movement_copy['amount'] = Decimal(movement_str_data['amount'])
            duplicates_for_template.append(movement_copy)

        return render(request, self.template_name, {
            'duplicates': duplicates_for_template
        })

    def post(self, request):
        pending_import = request.session.get('pending_import')
        if not pending_import:
            return redirect('bank_import')

        account = get_object_or_404(Account, id=pending_import['account_id'])
        session_duplicates_str = pending_import['duplicates']
        
        movements_to_process = []
        for movement_data in session_duplicates_str: # Iterate directly over the list of dicts
            movements_to_process.append(movement_data)

        selected_movements = []
        for movement in movements_to_process:
            # La clave para los datos POST debe coincidir con cómo se genera en la plantilla.
            # El diccionario 'movement' ya tiene 'date' y 'amount' como strings desde la sesión.
            if isinstance(movement['date'], str) and isinstance(movement['amount'], str):
                movement_key = f"{movement['date']}_{movement['amount']}"
                if request.POST.get(movement_key) == 'import':
                    selected_movements.append(movement)

        if selected_movements:
            success_count = self.import_movements(account, selected_movements)
            messages.success(request, f'Se han importado {success_count} movimientos duplicados seleccionados.')

        if 'pending_import' in request.session:
            del request.session['pending_import']
        return redirect('bank_import')
